from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from typing import List, Dict, Any
import io

# ── Helpers ───────────────────────────────────────────────────────────────────

def hex_fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor="FF" + hex6.upper())

def thin_border() -> Border:
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(cell, fill6: str, font6: str = "000000", bold: bool = True, align: str = "center"):
    cell.fill = hex_fill(fill6)
    cell.font = Font(bold=bold, color="FF" + font6.upper(), name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = thin_border()

# ── Platform config ───────────────────────────────────────────────────────────

PLATFORM_ORDER  = ["talabat", "keeta", "noon", "careem", "deliveroo", "smiles"]
PLATFORM_LABELS = {"talabat": "Talabat", "keeta": "Keeta", "noon": "Noon",
                   "careem": "Careem", "deliveroo": "Deliveroo", "smiles": "Smiles"}

PH_FILL = {"talabat": "FF9900", "keeta": "FBBC04", "noon": "4285F4",
           "careem": "34A853", "deliveroo": "46BDC6", "smiles": "EA4335"}
PH_FONT = {"talabat": "000000", "keeta": "000000", "noon": "FFFFFF",
           "careem": "FFFFFF", "deliveroo": "FFFFFF", "smiles": "FFFFFF"}
PD_FILL = {"talabat": "FCE5CD", "keeta": "FFF2CC", "noon": "C9DAF8",
           "careem": "D9EAD3", "deliveroo": "D0E0E3", "smiles": "F4CCCC"}

# ── KPI rows ──────────────────────────────────────────────────────────────────

KPI_ROWS = [
    ("اجمالي الطلبات",                     "Total Orders",         "numOrders",      "number"),
    ("اجمالي المبيعات",                    "Total Sales",          "totalSales",     "currency"),
    ("الخصم",                               "Discount",             "discount",       "currency"),
    ("المبلغ المحتسب منه العموله",          "Earnings",             "earnings",       "currency"),
    ("سعر المطعم",                          "Actual sales",         "actualSales",    "currency"),
    ("المصروفات",                           "Expenses",             "expenses",       "currency"),
    ("الارباح",                             "Net Revenue",          "netRevenue",     "currency"),
    ("الفرق بين الربح وسعر المطعم الاصلي", "Deffirence",           "difference",     "currency"),
    ("تكلفه الطعام",                        "Food Cost",            "foodCost",       "currency"),
    ("الفرق بين الربح والتكلفه",            "Deffirence Food Cost", "differenceCost", "currency"),
]

PCT_FORMULAS = [
    "1", "1",
    "=(J7/J6)", "=(J8/J6)", "=(J9/J6)", "=(J10/J6)",
    "=(K6+K12)", "=(J12/J11)", "=(J13/J11)", "=(J14/J9)",
]

# ── Main function ─────────────────────────────────────────────────────────────

def generate_master_sheet(
    restaurant_name: str,
    date_from: str,
    date_to: str,
    platform_results: List[Dict],
    total_kpi: Dict,
) -> bytes:

    selected    = [p for p in PLATFORM_ORDER if any(r["platform"] == p for r in platform_results)]
    result_map  = {r["platform"]: r["kpi"] for r in platform_results}

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, w in [("A", 0.38), ("B", 24), ("C", 19.5), ("D", 13.75), ("E", 13.75),
                   ("F", 13.75), ("G", 13.75), ("H", 13.75), ("I", 13.75),
                   ("J", 18), ("K", 13), ("L", 0.38)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 1.5

    # ── Row 2: Restaurant Name ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    ws["D2"] = "Restaurant Name:"
    sc(ws["D2"], "EAD1DC")
    ws.merge_cells("D2:E2")

    ws["F2"] = restaurant_name
    sc(ws["F2"], "674EA7", "FFFFFF")
    ws.merge_cells("F2:I2")

    ws.merge_cells("J2:K3")
    sc(ws["J2"], "A64D79", "000000", False, "left")

    # ── Row 3: Period ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    ws["D3"] = "Period:"
    sc(ws["D3"], "EAD1DC")
    ws.merge_cells("D3:E3")

    ws["F3"] = "From:"
    sc(ws["F3"], "EAD1DC")

    ws["G3"] = date_from
    sc(ws["G3"], "674EA7", "FFFFFF")

    ws["H3"] = "To:"
    sc(ws["H3"], "EAD1DC")

    ws["I3"] = date_to
    sc(ws["I3"], "674EA7", "FFFFFF")

    # ── Row 4: Platform headers ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws["B4"] = "المنصات"
    sc(ws["B4"], "A64D79", "FFFFFF", True, "right")

    ws["C4"] = "Aggregator"
    sc(ws["C4"], "A64D79", "FFFFFF")

    platform_cols = {}
    for i, p in enumerate(selected):
        col = get_column_letter(4 + i)
        platform_cols[p] = col
        ws[f"{col}4"] = PLATFORM_LABELS[p]
        sc(ws[f"{col}4"], PH_FILL[p], PH_FONT[p])

    total_col = get_column_letter(4 + len(selected))
    pct_col   = get_column_letter(4 + len(selected) + 1)

    ws[f"{total_col}4"] = "Total"
    sc(ws[f"{total_col}4"], "A64D79", "FFFFFF")

    ws[f"{pct_col}4"] = "Percent"
    sc(ws[f"{pct_col}4"], "A64D79", "FFFFFF")

    # ── Data rows 5-14 ────────────────────────────────────────────────────────
    for row_i, (ar, en, key, fmt) in enumerate(KPI_ROWS):
        r = 5 + row_i
        ws.row_dimensions[r].height = 20

        ws[f"B{r}"] = ar
        sc(ws[f"B{r}"], "EAD1DC", "000000", True, "right")

        ws[f"C{r}"] = en
        sc(ws[f"C{r}"], "EAD1DC", "000000", True, "left")

        for p in selected:
            col = platform_cols[p]
            val = round(result_map.get(p, {}).get(key, 0), 2)
            ws[f"{col}{r}"] = val
            sc(ws[f"{col}{r}"], PD_FILL[p])
            ws[f"{col}{r}"].number_format = "#,##0.00" if fmt == "currency" else "#,##0"

        total_val = round(total_kpi.get(key, 0), 2)
        ws[f"{total_col}{r}"] = total_val
        sc(ws[f"{total_col}{r}"], "EAD1DC")
        ws[f"{total_col}{r}"].number_format = "#,##0.00" if fmt == "currency" else "#,##0"

        pct = PCT_FORMULAS[row_i]
        if pct == "1":
            ws[f"{pct_col}{r}"] = 1.0
            ws[f"{pct_col}{r}"].number_format = "0.00%"
        else:
            ws[f"{pct_col}{r}"] = pct
            ws[f"{pct_col}{r}"].number_format = "0.0%"
        sc(ws[f"{pct_col}{r}"], "674EA7", "FFFFFF")

    # ── Merge spacer ──────────────────────────────────────────────────────────
    ws.merge_cells("B1:C3")
    ws["B1"].fill = hex_fill("A64D79")

    # ── Return as bytes ───────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()